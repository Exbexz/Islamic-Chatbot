from openpyxl import load_workbook


def getTotalQuestions(path):

	workbook = load_workbook(path) 
	sheet = workbook['Sheet1'] 
	num_rows = sheet.max_row
	workbook.close()

	return num_rows


def getQuestion(path,index):

	workbook = load_workbook(path) 
	sheet = workbook['Sheet1'] 
	question = sheet.cell(row=index,column=1)
	workbook.close()

	return question.value


def getAnswer(path,index):

	workbook = load_workbook(path) 
	sheet = workbook['Sheet1'] 
	answer = sheet.cell(row=index,column=2)
	workbook.close()

	return answer.value


